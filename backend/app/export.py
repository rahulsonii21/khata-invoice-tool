"""
Generates PDF statements and Excel exports for parties and invoices.

PDF: formatted statement per party (invoice list + payment history + balance).
Excel: raw data dump, filterable, either for one party or all parties at once.
"""
import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.lib.enums import TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INK = colors.HexColor("#1F3A34")
RUST = colors.HexColor("#A13D2B")
SAGE = colors.HexColor("#DCE5DE")
LINE = colors.HexColor("#D8D2C0")


def _fmt_inr(amount):
    if amount is None:
        return "-"
    return f"Rs {amount:,.2f}"


# ---------------------------------------------------------------------------
# PDF - single party statement
# ---------------------------------------------------------------------------
def generate_party_statement_pdf(party, invoices, company_settings=None, logo_bytes=None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm,
        leftMargin=18 * mm, rightMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"], textColor=INK, fontSize=18, spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], textColor=colors.HexColor("#7C9089"), fontSize=10,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], textColor=INK, fontSize=12, spaceBefore=14, spaceAfter=6,
    )
    letterhead_style = ParagraphStyle(
        "Letterhead", parent=styles["Normal"], textColor=INK, fontSize=9, alignment=TA_RIGHT,
    )

    elements = []

    # Company letterhead, right-aligned at the top, if the user has set it up
    if company_settings and company_settings.company_name:
        if logo_bytes:
            try:
                logo_img = RLImage(io.BytesIO(logo_bytes), width=20 * mm, height=20 * mm)
                logo_img.hAlign = "RIGHT"
                elements.append(logo_img)
                elements.append(Spacer(1, 2 * mm))
            except Exception:
                pass  # a corrupt/unreadable logo file shouldn't break the whole export
        lh_lines = [f"<b>{company_settings.company_name}</b>"]
        if company_settings.gstin:
            lh_lines.append(f"GSTIN: {company_settings.gstin}")
        if company_settings.address:
            lh_lines.append(company_settings.address)
        if company_settings.phone:
            lh_lines.append(company_settings.phone)
        elements.append(Paragraph("<br/>".join(lh_lines), letterhead_style))
        elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("Account Statement", title_style))
    elements.append(Paragraph(f"{party.name}", subtitle_style))
    if party.phone:
        elements.append(Paragraph(f"Phone: {party.phone}", subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y')}", subtitle_style))
    elements.append(Spacer(1, 10 * mm))

    # Summary box
    total_invoiced = sum(inv.amount for inv in invoices)
    total_received = sum(p.amount for inv in invoices for p in inv.payments)
    outstanding = total_invoiced - total_received

    summary_data = [
        ["Total Invoiced", "Total Received", "Outstanding"],
        [_fmt_inr(total_invoiced), _fmt_inr(total_received), _fmt_inr(outstanding)],
    ]
    summary_table = Table(summary_data, colWidths=[52 * mm, 52 * mm, 52 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SAGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), INK),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("TEXTCOLOR", (2, 1), (2, 1), RUST if outstanding > 0 else INK),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
    ]))
    elements.append(summary_table)

    # Per-invoice breakdown
    elements.append(Paragraph("Invoices & Payments", section_style))

    for inv in invoices:
        inv_header = [
            f"Invoice {inv.invoice_number or '(no number)'} — "
            f"{inv.invoice_date.strftime('%d %b %Y') if inv.invoice_date else 'no date'}",
            _fmt_inr(inv.amount),
            inv.status.value.replace("_", " ").title(),
        ]
        rows = [inv_header]
        for p in inv.payments:
            rows.append([
                f"    Payment — {p.payment_date.strftime('%d %b %Y')} ({p.mode.value})"
                + (f" — {p.remarks}" if p.remarks else ""),
                _fmt_inr(-p.amount),
                "",
            ])
        rows.append(["    Outstanding", _fmt_inr(inv.outstanding), ""])

        t = Table(rows, colWidths=[98 * mm, 35 * mm, 23 * mm])
        style = [
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (0, 0), INK),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, LINE),
            ("LINEBELOW", (0, -1), (-1, -1), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TEXTCOLOR", (0, -1), (-1, -1), RUST if inv.outstanding > 0 else INK),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        t.setStyle(TableStyle(style))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

    if not invoices:
        elements.append(Paragraph("No invoices on record for this party.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF - monthly summary across parties (totals only, one row per party)
# ---------------------------------------------------------------------------
def generate_monthly_summary_pdf(month_label, party_invoice_pairs, company_settings=None, logo_bytes=None) -> bytes:
    """
    party_invoice_pairs: list of (party, invoices) tuples, invoices already
    filtered to the month being reported on.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm,
        leftMargin=18 * mm, rightMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=INK, fontSize=18, spaceAfter=2)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], textColor=colors.HexColor("#7C9089"), fontSize=10)
    letterhead_style = ParagraphStyle("Letterhead", parent=styles["Normal"], textColor=INK, fontSize=9, alignment=TA_RIGHT)

    elements = []

    if company_settings and company_settings.company_name:
        if logo_bytes:
            try:
                logo_img = RLImage(io.BytesIO(logo_bytes), width=20 * mm, height=20 * mm)
                logo_img.hAlign = "RIGHT"
                elements.append(logo_img)
                elements.append(Spacer(1, 2 * mm))
            except Exception:
                pass
        elements.append(Paragraph(f"<b>{company_settings.company_name}</b>", letterhead_style))
        elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph(f"Monthly Sales Report — {month_label}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y')}", subtitle_style))
    elements.append(Spacer(1, 8 * mm))

    grand_invoiced = sum(sum(inv.amount for inv in invs) for _, invs in party_invoice_pairs)
    grand_received = sum(sum(p.amount for inv in invs for p in inv.payments) for _, invs in party_invoice_pairs)

    summary_data = [["Total Invoiced", "Total Received", "Outstanding (this month's bills)"]]
    summary_data.append([_fmt_inr(grand_invoiced), _fmt_inr(grand_received), _fmt_inr(grand_invoiced - grand_received)])
    summary_table = Table(summary_data, colWidths=[52 * mm, 52 * mm, 52 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SAGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), INK),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8 * mm))

    # Per-party breakdown table
    rows = [["Party", "Invoiced", "Received", "Outstanding", "# Bills"]]
    for party, invs in sorted(party_invoice_pairs, key=lambda x: sum(i.amount for i in x[1]), reverse=True):
        if not invs:
            continue
        invoiced = sum(i.amount for i in invs)
        received = sum(p.amount for i in invs for p in i.payments)
        rows.append([party.name, _fmt_inr(invoiced), _fmt_inr(received), _fmt_inr(invoiced - received), str(len(invs))])

    if len(rows) > 1:
        t = Table(rows, colWidths=[62 * mm, 32 * mm, 32 * mm, 32 * mm, 18 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SAGE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No invoices in this period.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF - combined bill images, one per page, chronological, with captions
# ---------------------------------------------------------------------------
def generate_combined_bills_pdf(invoices_with_images) -> bytes:
    """
    invoices_with_images: list of (invoice, party_name, image_bytes) tuples,
    already sorted into the desired page order by the caller.
    Skips any entry where image_bytes is None (missing/unreadable image)
    rather than failing the whole export.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    caption_style = ParagraphStyle(
        "Caption", parent=styles["Normal"], textColor=INK, fontSize=10, spaceAfter=4,
    )

    elements = []
    included = 0
    skipped = 0

    for invoice, party_name, image_bytes in invoices_with_images:
        if not image_bytes:
            skipped += 1
            continue
        try:
            date_str = invoice.invoice_date.strftime("%d %b %Y") if invoice.invoice_date else "no date"
            caption = f"<b>{party_name}</b> — Invoice {invoice.invoice_number or '(no number)'} — {date_str} — {_fmt_inr(invoice.amount)}"
            elements.append(Paragraph(caption, caption_style))

            img = RLImage(io.BytesIO(image_bytes))
            # Fit within the page while preserving aspect ratio
            max_w, max_h = 180 * mm, 240 * mm
            ratio = min(max_w / img.imageWidth, max_h / img.imageHeight, 1)
            img.drawWidth = img.imageWidth * ratio
            img.drawHeight = img.imageHeight * ratio
            elements.append(img)
            elements.append(Spacer(1, 10 * mm))
            included += 1
        except Exception:
            skipped += 1
            continue

    if included == 0:
        elements.append(Paragraph("No bill images found for this period.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()
HEADER_FILL = PatternFill(start_color="1F3A34", end_color="1F3A34", fill_type="solid")
HEADER_FONT = Font(color="FBFAF6", bold=True)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def generate_excel_export(parties_with_invoices) -> bytes:
    """
    parties_with_invoices: list of (party, invoices) tuples.
    Produces a workbook with:
      - Summary sheet (one row per party)
      - Invoices sheet (one row per invoice)
      - Payments sheet (one row per payment)
    """
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Party", "Phone", "Total Invoiced", "Total Received", "Outstanding"])
    for party, invoices in parties_with_invoices:
        total_invoiced = sum(inv.amount for inv in invoices)
        total_received = sum(p.amount for inv in invoices for p in inv.payments)
        ws_summary.append([
            party.name, party.phone or "", total_invoiced, total_received,
            total_invoiced - total_received,
        ])
    _style_header(ws_summary)
    _autosize(ws_summary)

    # Invoices sheet
    ws_inv = wb.create_sheet("Invoices")
    ws_inv.append([
        "Party", "Invoice Number", "Invoice Date", "Amount", "GST Amount",
        "Total Paid", "Outstanding", "Status", "Remarks",
    ])
    for party, invoices in parties_with_invoices:
        for inv in invoices:
            ws_inv.append([
                party.name, inv.invoice_number or "",
                inv.invoice_date.isoformat() if inv.invoice_date else "",
                inv.amount, inv.gst_amount or 0,
                inv.total_paid, inv.outstanding,
                inv.status.value.replace("_", " ").title(),
                inv.remarks or "",
            ])
    _style_header(ws_inv)
    _autosize(ws_inv)

    # Payments sheet
    ws_pay = wb.create_sheet("Payments")
    ws_pay.append([
        "Party", "Invoice Number", "Payment Date", "Amount", "Mode", "Remarks",
    ])
    for party, invoices in parties_with_invoices:
        for inv in invoices:
            for p in inv.payments:
                ws_pay.append([
                    party.name, inv.invoice_number or "",
                    p.payment_date.isoformat(), p.amount, p.mode.value, p.remarks or "",
                ])
    _style_header(ws_pay)
    _autosize(ws_pay)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Purchase ledger (payables) reports - mirror the sales-side functions above,
# with the correct labels/attribute names for Supplier/Purchase objects
# rather than Party/Invoice. The combined-bills-PDF function above is reused
# as-is via a small adapter in the router (it's generic enough already).
# ---------------------------------------------------------------------------
def generate_monthly_purchase_summary_pdf(month_label, supplier_purchase_pairs, company_settings=None, logo_bytes=None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm,
        leftMargin=18 * mm, rightMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=INK, fontSize=18, spaceAfter=2)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], textColor=colors.HexColor("#7C9089"), fontSize=10)
    letterhead_style = ParagraphStyle("Letterhead", parent=styles["Normal"], textColor=INK, fontSize=9, alignment=TA_RIGHT)

    elements = []

    if company_settings and company_settings.company_name:
        if logo_bytes:
            try:
                logo_img = RLImage(io.BytesIO(logo_bytes), width=20 * mm, height=20 * mm)
                logo_img.hAlign = "RIGHT"
                elements.append(logo_img)
                elements.append(Spacer(1, 2 * mm))
            except Exception:
                pass
        elements.append(Paragraph(f"<b>{company_settings.company_name}</b>", letterhead_style))
        elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph(f"Monthly Purchase Report — {month_label}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y')}", subtitle_style))
    elements.append(Spacer(1, 8 * mm))

    grand_purchased = sum(sum(pu.amount for pu in purs) for _, purs in supplier_purchase_pairs)
    grand_paid = sum(sum(p.amount for pu in purs for p in pu.payments) for _, purs in supplier_purchase_pairs)

    summary_data = [["Total Purchased", "Total Paid", "Payable (this month's bills)"]]
    summary_data.append([_fmt_inr(grand_purchased), _fmt_inr(grand_paid), _fmt_inr(grand_purchased - grand_paid)])
    summary_table = Table(summary_data, colWidths=[52 * mm, 52 * mm, 52 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SAGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), INK),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8 * mm))

    rows = [["Supplier", "Purchased", "Paid", "Payable", "# Bills"]]
    for supplier, purs in sorted(supplier_purchase_pairs, key=lambda x: sum(p.amount for p in x[1]), reverse=True):
        if not purs:
            continue
        purchased = sum(p.amount for p in purs)
        paid = sum(pay.amount for p in purs for pay in p.payments)
        rows.append([supplier.name, _fmt_inr(purchased), _fmt_inr(paid), _fmt_inr(purchased - paid), str(len(purs))])

    if len(rows) > 1:
        t = Table(rows, colWidths=[62 * mm, 32 * mm, 32 * mm, 32 * mm, 18 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SAGE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No purchases in this period.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


def generate_purchase_excel_export(suppliers_with_purchases) -> bytes:
    """
    suppliers_with_purchases: list of (supplier, purchases) tuples.
    Same structure as generate_excel_export, for the payables side.
    """
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Supplier", "Phone", "Total Purchased", "Total Paid", "Payable"])
    for supplier, purchases in suppliers_with_purchases:
        total_purchased = sum(p.amount for p in purchases)
        total_paid = sum(pay.amount for p in purchases for pay in p.payments)
        ws_summary.append([
            supplier.name, supplier.phone or "", total_purchased, total_paid,
            total_purchased - total_paid,
        ])
    _style_header(ws_summary)
    _autosize(ws_summary)

    ws_pur = wb.create_sheet("Purchases")
    ws_pur.append([
        "Supplier", "Bill Number", "Purchase Date", "Amount", "GST Amount",
        "Total Paid", "Payable", "Status", "Remarks",
    ])
    for supplier, purchases in suppliers_with_purchases:
        for pu in purchases:
            ws_pur.append([
                supplier.name, pu.purchase_number or "",
                pu.purchase_date.isoformat() if pu.purchase_date else "",
                pu.amount, pu.gst_amount or 0,
                pu.total_paid, pu.outstanding,
                pu.status.value.replace("_", " ").title(),
                pu.remarks or "",
            ])
    _style_header(ws_pur)
    _autosize(ws_pur)

    ws_pay = wb.create_sheet("Payments")
    ws_pay.append(["Supplier", "Bill Number", "Payment Date", "Amount", "Mode", "Remarks"])
    for supplier, purchases in suppliers_with_purchases:
        for pu in purchases:
            for p in pu.payments:
                ws_pay.append([
                    supplier.name, pu.purchase_number or "",
                    p.payment_date.isoformat(), p.amount, p.mode.value, p.remarks or "",
                ])
    _style_header(ws_pay)
    _autosize(ws_pay)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
